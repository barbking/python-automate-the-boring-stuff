import os
import openpyxl

os.chdir('c:\\Users\\barba\\Documents\\Automate_the_Boring_Stuff_2e_onlinematerials\\automate_online-materials')
workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook.get_sheet_by_name('Sheet1')
print(workbook.get_sheet_names())

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1'] # Get a sheet from the workbook.
sheet['A1'] # Get a cell from the sheet.
#<Cell 'Sheet1'.A1>
sheet['A1'].value # Get the value from the cell.
#datetime.datetime(2015, 4, 5, 13, 34, 2)
c = sheet['B1'] # Get another cell from the sheet.
c.value
#'Apples'
# Get the row, column, and value from the cell.
'Row %s, Column %s is %s' % (c.row, c.column, c.value)
#'Row 1, Column B is Apples'
'Cell %s is %s' % (c.coordinate, c.value)
#'Cell B1 is Apples'
sheet['C1']

#getting rows and cols from sheet
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
tuple(sheet['A1':'C3']) # Get all cells from A1 to C3.
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

sheet = wb.active
list(sheet.columns)[1] # Get second column's cells.
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)